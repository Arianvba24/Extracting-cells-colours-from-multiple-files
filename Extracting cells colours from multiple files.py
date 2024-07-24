#Take all the excel files of the folder C:\Users\Cash\Documents\pruebas_python\proyectos\excel
#merge the data of the sheets 'Stocks' creating a dataframe and extract the definition of each
# color
# Red = "Loss"
# Yellow = "Damaged"
# Green = "Sold"
# White = "Not reviewed"

# ----------------------------------------------------------------------------
# Colour values:
#Verde: FF92D050
#Rojo: FFFF0000
#Amarillo: FFFFFF00
#Blanco: 00000000

import os

files = os.listdir(r"C:\Users\Cash\Documents\pruebas_python\proyectos\excel")

# files = os.listdir("./excel_files")

colours = {"FF92D050": "Sold", "FFFF0000": "Loss", "FFFFFF00": "Damaged", "FFFFFFFF" : "Not reviewed"}
import openpyxl

bar = "\\"
for file in files:
    file_path = fr"C:\Users\Cash\Documents\pruebas_python\proyectos\excel/{file}"
    excel = openpyxl.load_workbook(file_path)

    ws= excel.active


    rows = 0
    for i in range(2,1001):
        if ws[f"A{i}"].value is not None:
            rows = rows + 1

        else:
            pass

    for row in range(2,rows+2):
        ws[f"E{row}"].value = colours.get(ws[f"D{row}"].fill.start_color.index)


    excel.save(file_path)

import pandas as pd


df = pd.DataFrame()

for file in files:
    file_path = fr"C:\Users\Cash\Documents\pruebas_python\proyectos\excel/{file}"
    # file_path = fr".\excel_files/{file}"
    df_x = pd.read_excel(file_path, usecols="A:E")
    df = pd.concat([df,df_x], ignore_index = True)
    
df.to_excel(r"C:\Users\Cash\Proyectos\Excel\dataframe.xlsx",index=False)
print(df)

